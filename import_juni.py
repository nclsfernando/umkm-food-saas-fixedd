import openpyxl, json, requests, sys
from pathlib import Path
from datetime import datetime

# ============================================
# Jalankan: python3 import_juni.py <folder_files>
# Contoh:   python3 import_juni.py ~/Downloads
# ============================================

API = 'https://umkm-food-saas-fixedd-production.up.railway.app/api/v1'
EMAIL = 'admin@umkm.id'
PASSWORD = 'felice123'

folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('.')
files = sorted(folder.glob('GrabMerchant_Reports_*.xlsx'))
print(f"Ditemukan {len(files)} file GrabMerchant")

# Login
login = requests.post(f'{API}/auth/login', json={'email': EMAIL, 'password': PASSWORD})
token = login.json().get('access_token')
if not token:
    print("Login gagal:", login.text)
    sys.exit(1)
print(f"Login OK")

headers_req = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
all_orders = []

for f in files:
    wb = openpyxl.load_workbook(f)
    if 'Transaksi' not in wb.sheetnames:
        print(f"  Skip {f.name} - sheet Transaksi tidak ditemukan")
        continue
    ws = wb['Transaksi']
    hdrs = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(hdrs) if h}
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        kat = str(row[col.get('Kategori', 999)] or '').strip()
        if kat != 'Pembayaran': continue
        jumlah = float(row[col.get('Jumlah', 999)] or 0)
        if jumlah <= 0: continue
        
        def g(key, default=0):
            idx = col.get(key)
            return row[idx] if idx is not None else default
        
        biaya_jasa = abs(float(g('Biaya jasa') or 0))
        biaya_sukses = abs(float(g('Biaya sukses pemasaran') or 0))
        mdr = abs(float(g('Nilai MDR bersih') or 0))
        commission = biaya_jasa + biaya_sukses + mdr
        total_val = float(g('Total') or 0)
        
        tgl_raw = g('Tanggal dibuat', None) or g('Diperbarui Pada', None)
        tgl_str = tgl_raw.isoformat() if isinstance(tgl_raw, datetime) else str(tgl_raw or datetime.now())
        
        item_meta = json.dumps({
            'jenis': str(g('Jenis', '') or ''),
            'metode': str(g('Metode pembayaran', '') or ''),
            'idPesanan': str(g('ID pesanan pendek', '') or g('ID transaksi', '') or ''),
            'biayaJasa': biaya_jasa,
            'biayaSukses': biaya_sukses,
            'mdr': mdr,
            'tanggalTransfer': str(g('Tanggal transfer', '') or ''),
            'idPencairan': str(g('ID pencairan dana', '') or ''),
        })
        
        all_orders.append({
            'orderDate': tgl_str,
            'marketplace': 'GrabFood',
            'grossSales': jumlah,
            'discount': 0,
            'commission': commission,
            'netSales': total_val if total_val > 0 else jumlah - commission,
            'status': 'COMPLETED',
            'items': [{'productName': item_meta, 'qty': 1, 'unitPrice': jumlah, 'subtotal': jumlah}]
        })
        count += 1
    print(f"  {f.name}: {count} transaksi")

print(f"\nTotal: {len(all_orders)} transaksi, mengimport...")
created = 0
for order in all_orders:
    res = requests.post(f'{API}/orders', json=order, headers=headers_req)
    if res.status_code in [200, 201]:
        created += 1
    else:
        print(f"  Error: {res.status_code} - {res.text[:80]}")

print(f"\nSelesai! {created}/{len(all_orders)} transaksi berhasil diimport.")
